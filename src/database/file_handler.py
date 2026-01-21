import csv
import json
import logging
import os
from datetime import datetime
from typing import Any, Dict, List, Optional

import yaml

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from ..models.dataset import Dataset
from ..models.experiment import Experiment
from ..models.result import Result

# Configure logging
logger = logging.getLogger(__name__)


class FileHandler:
    """
    File Handler - manages file I/O operations.
    """

    def __init__(self, output_dir: str = "exports"):
        """
        Initialize the file handler.
        """
        self._output_dir = output_dir
        self._ensure_directory(output_dir)

    @staticmethod
    def _ensure_directory(path: str) -> None:
        """Create directory if it doesn't exist."""
        if path and not os.path.exists(path):
            os.makedirs(path)
            logger.info(f"Created directory: {path}")

    def _get_filepath(self, filename: str) -> str:
        """Get full filepath in output directory."""
        return os.path.join(self._output_dir, filename)

    @staticmethod
    def load_yaml(filepath: str) -> Dict[str, Any]:
        """
        Load configuration from a YAML file.
        """
        try:
            with open(filepath, "r", encoding="utf-8") as f:
                data = yaml.safe_load(f)
            logger.info(f"Loaded YAML config from: {filepath}")
            return data or {}
        except FileNotFoundError:
            logger.error(f"YAML file not found: {filepath}")
            return {}
        except yaml.YAMLError as e:
            logger.error(f"Error parsing YAML: {e}")
            return {}

    @staticmethod
    def save_yaml(data: Dict[str, Any], filepath: str) -> bool:
        """
        Save data to a YAML file.

        Args:
            data: Dictionary to save
            filepath: Output file path

        Returns:
            bool: True if successful
        """
        try:
            with open(filepath, "w", encoding="utf-8") as f:
                yaml.dump(
                    data,
                    f,
                    default_flow_style=False,
                    allow_unicode=True,
                    sort_keys=False,
                )
            logger.info(f"Saved YAML to: {filepath}")
            return True
        except Exception as e:
            logger.error(f"Error saving YAML: {e}")
            return False

    # ==================== JSON OPERATIONS ====================

    def export_to_json(
        self,
        experiments: List[Experiment],
        filename: str = "experiments.json",
        include_metadata: bool = True,
    ) -> str:
        """
        Export experiments to JSON file.

        Args:
            experiments: List of experiments to export
            filename: Output filename
            include_metadata: Include export metadata

        Returns:
            str: Path to the created file
        """
        filepath = self._get_filepath(filename)

        data = {"experiments": [exp.to_dict() for exp in experiments]}

        if include_metadata:
            data["metadata"] = {
                "exported_at": datetime.now().isoformat(),
                "total_count": len(experiments),
                "format_version": "1.0",
            }

        try:
            with open(filepath, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
            logger.info(f"Exported {len(experiments)} experiments to JSON: {filepath}")
            return filepath
        except Exception as e:
            logger.error(f"Error exporting to JSON: {e}")
            raise

    def import_from_json(self, filepath: str) -> List[Experiment]:
        """
        Import experiments from JSON file.

        Args:
            filepath: Path to JSON file

        Returns:
            List[Experiment]: Imported experiments
        """
        try:
            with open(filepath, "r", encoding="utf-8") as f:
                data = json.load(f)

            experiments = []
            for exp_data in data.get("experiments", []):
                experiments.append(Experiment.from_dict(exp_data))

            logger.info(f"Imported {len(experiments)} experiments from JSON")
            return experiments
        except FileNotFoundError:
            logger.error(f"JSON file not found: {filepath}")
            return []
        except json.JSONDecodeError as e:
            logger.error(f"Invalid JSON format: {e}")
            return []

    # ==================== CSV OPERATIONS ====================

    def export_to_csv(
        self, experiments: List[Experiment], filename: str = "experiments.csv"
    ) -> str:
        """
        Export experiments to CSV file.

        Args:
            experiments: List of experiments to export
            filename: Output filename

        Returns:
            str: Path to the created file
        """
        filepath = self._get_filepath(filename)

        # Define CSV columns
        fieldnames = [
            "id",
            "name",
            "description",
            "author",
            "status",
            "tags",
            "created_at",
            "updated_at",
            "accuracy",
            "loss",
            "f1_score",
        ]

        try:
            with open(filepath, "w", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()

                for exp in experiments:
                    row = {
                        "id": exp.id,
                        "name": exp.name,
                        "description": exp.description,
                        "author": exp.author,
                        "status": exp.status.value,
                        "tags": "; ".join(exp.tags),
                        "created_at": exp.created_at.isoformat(),
                        "updated_at": exp.updated_at.isoformat(),
                        "accuracy": exp.metrics.get("accuracy", ""),
                        "loss": exp.metrics.get("loss", ""),
                        "f1_score": exp.metrics.get("f1_score", ""),
                    }
                    writer.writerow(row)

            logger.info(f"Exported {len(experiments)} experiments to CSV: {filepath}")
            return filepath
        except Exception as e:
            logger.error(f"Error exporting to CSV: {e}")
            raise

    def import_from_csv(self, filepath: str) -> List[Dict[str, Any]]:
        """
        Import data from CSV file.

        Args:
            filepath: Path to CSV file

        Returns:
            List[Dict]: List of row dictionaries
        """
        try:
            with open(filepath, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                data = list(reader)

            logger.info(f"Imported {len(data)} rows from CSV")
            return data
        except FileNotFoundError:
            logger.error(f"CSV file not found: {filepath}")
            return []
        except Exception as e:
            logger.error(f"Error reading CSV: {e}")
            return []

    # ==================== EXCEL OPERATIONS ====================

    def export_to_excel(
        self,
        experiments: List[Experiment],
        filename: str = "experiments.xlsx",
        sheet_name: str = "Experiments",
    ) -> str:
        """
        Export experiments to Excel file with formatting.

        Args:
            experiments: List of experiments to export
            filename: Output filename
            sheet_name: Name of the worksheet

        Returns:
            str: Path to the created file
        """
        if not OPENPYXL_AVAILABLE:
            logger.error("openpyxl not installed. Cannot export to Excel.")
            raise ImportError("openpyxl is required for Excel export")

        filepath = self._get_filepath(filename)

        # Create workbook and sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Define headers
        headers = [
            "ID",
            "Name",
            "Description",
            "Author",
            "Status",
            "Tags",
            "Created At",
            "Updated At",
            "Accuracy",
            "Loss",
            "F1 Score",
        ]

        # Style definitions
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_alignment = Alignment(horizontal="center", vertical="center")

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Write headers with styling
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Write data rows
        for row_idx, exp in enumerate(experiments, 2):
            data_row = [
                exp.id,
                exp.name,
                exp.description,
                exp.author,
                exp.status.value,
                "; ".join(exp.tags),
                exp.created_at.strftime("%Y-%m-%d %H:%M"),
                exp.updated_at.strftime("%Y-%m-%d %H:%M"),
                exp.metrics.get("accuracy", ""),
                exp.metrics.get("loss", ""),
                exp.metrics.get("f1_score", ""),
            ]

            for col, value in enumerate(data_row, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border

                # Alternate row colors
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(
                        start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"
                    )

        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            max_length = 0
            column = get_column_letter(col)

            for cell in ws[column]:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        # Freeze the header row
        ws.freeze_panes = "A2"

        # Save the workbook
        try:
            wb.save(filepath)
            logger.info(f"Exported {len(experiments)} experiments to Excel: {filepath}")
            return filepath
        except Exception as e:
            logger.error(f"Error saving Excel file: {e}")
            raise
        finally:
            wb.close()

    # ==================== REPORT GENERATION ====================

    def generate_summary_report(
        self, experiments: List[Experiment], filename: str = "summary_report.xlsx"
    ) -> str:
        """
        Generate a comprehensive summary report in Excel format.

        Args:
            experiments: List of experiments
            filename: Output filename

        Returns:
            str: Path to the created file
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for report generation")

        filepath = self._get_filepath(filename)
        wb = openpyxl.Workbook()

        # Summary Sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"

        # Calculate statistics
        total = len(experiments)
        by_status = {}
        for exp in experiments:
            status = exp.status.value
            by_status[status] = by_status.get(status, 0) + 1

        # Write summary
        ws_summary["A1"] = "Research Experiments Summary Report"
        ws_summary["A1"].font = Font(bold=True, size=16)

        ws_summary["A3"] = "Generated At:"
        ws_summary["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        ws_summary["A4"] = "Total Experiments:"
        ws_summary["B4"] = total

        row = 6
        ws_summary[f"A{row}"] = "Status Breakdown:"
        ws_summary[f"A{row}"].font = Font(bold=True)

        for status, count in by_status.items():
            row += 1
            ws_summary[f"A{row}"] = f"  {status.capitalize()}:"
            ws_summary[f"B{row}"] = count

        # Details Sheet
        ws_details = wb.create_sheet("Experiments")
        headers = ["Name", "Author", "Status", "Created", "Tags"]

        for col, header in enumerate(headers, 1):
            cell = ws_details.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)

        for row_idx, exp in enumerate(experiments, 2):
            ws_details.cell(row=row_idx, column=1, value=exp.name)
            ws_details.cell(row=row_idx, column=2, value=exp.author)
            ws_details.cell(row=row_idx, column=3, value=exp.status.value)
            ws_details.cell(
                row=row_idx, column=4, value=exp.created_at.strftime("%Y-%m-%d")
            )
            ws_details.cell(row=row_idx, column=5, value="; ".join(exp.tags))

        wb.save(filepath)
        wb.close()

        logger.info(f"Generated summary report: {filepath}")
        return filepath

    # ==================== UTILITY METHODS ====================

    def export_all_formats(
        self, experiments: List[Experiment], base_filename: str = "experiments"
    ) -> Dict[str, str]:
        """
        Export experiments to all supported formats.

        Args:
            experiments: List of experiments
            base_filename: Base name for output files

        Returns:
            Dict[str, str]: Dictionary of format to filepath
        """
        results = {}

        # JSON
        results["json"] = self.export_to_json(experiments, f"{base_filename}.json")

        # CSV
        results["csv"] = self.export_to_csv(experiments, f"{base_filename}.csv")

        # Excel
        if OPENPYXL_AVAILABLE:
            results["xlsx"] = self.export_to_excel(experiments, f"{base_filename}.xlsx")

        return results
